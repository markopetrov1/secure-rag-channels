"""Build the human annotation package from the blinded sheet.

Two annotators grade the same items so that inter-annotator agreement can be
reported as the ceiling any automated judge is being measured against. Each gets
the items in a different seeded order, because grading is done over a session and
a shared order would let fatigue and drift correlate between the two annotators,
which is exactly the correlation the agreement statistic is supposed to detect.

Produces, per annotator, a spreadsheet with a constrained verdict column and a
plain CSV for anyone who would rather work in Google Sheets, plus a README.

Usage:
  python src/make_annotation_pack.py
  python src/make_annotation_pack.py --annotators "Marko Petrov,Ema Pandilova"
"""
import argparse
import csv
import random
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
HV = ROOT / "results/human_validation"
OUT = HV / "pack"

VERDICTS = ["correct", "missing", "incorrect", "unsure"]

GUIDE = """Grade the candidate answer against the reference material only.

  correct    every substantive claim in the answer agrees with the reference and
             the answer addresses the question that was asked
  missing    the answer declines, says it cannot answer, or leaves out the key
             information the question asks for
  incorrect  the answer makes at least one substantive claim that contradicts
             the reference, or invents a fact the reference does not support
  unsure     you genuinely cannot decide from the reference given

Length and style earn nothing. A one-line answer that is right is correct; a
polished paragraph that is wrong is incorrect. If the answer is right about the
substance but adds a claim the reference contradicts, that is incorrect.

Use unsure sparingly, and only where the reference is inadequate rather than
where the answer is hard to call. A forced guess adds noise; an honest unsure
does not, because these items are reported separately.

Do not open judge_key_DO_NOT_OPEN_BEFORE_ANNOTATING.csv until you have finished.
The whole point of this exercise is to compare your labels against the automated
ones, and seeing them first destroys that.
"""


def slug(name):
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def load_items():
    with (HV / "annotation_sheet.csv").open() as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, annotator):
    cols = ["row", "item_id", "question", "reference", "candidate_answer",
            "human_verdict", "human_notes"]
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for i, r in enumerate(rows, 1):
            w.writerow({"row": i, "item_id": r["item_id"],
                        "question": r["question"], "reference": r["reference"],
                        "candidate_answer": r["candidate_answer"],
                        "human_verdict": "", "human_notes": ""})


def write_xlsx(path, rows, annotator):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # Instructions first, so the file opens on them rather than on row one.
    ws0 = wb.active
    ws0.title = "How to grade"
    ws0["A1"] = f"Annotation sheet for {annotator}"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A3"] = (f"{len(rows)} items on the 'Grade' tab. Fill the "
                 f"human_verdict column. It is a dropdown, so you cannot mistype.")
    ws0["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws0.merge_cells("A3:F4")
    for i, line in enumerate(GUIDE.split("\n"), start=6):
        ws0.cell(row=i, column=1, value=line)
    ws0.column_dimensions["A"].width = 100

    ws = wb.create_sheet("Grade")
    head = ["row", "item_id", "question", "reference", "candidate_answer",
            "human_verdict", "human_notes"]
    ws.append(head)
    fill = PatternFill("solid", fgColor="DDE7F5")
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    for i, r in enumerate(rows, 1):
        ws.append([i, r["item_id"], r["question"], r["reference"],
                   r["candidate_answer"], "", ""])

    widths = {"A": 6, "B": 10, "C": 52, "D": 62, "E": 62, "F": 16, "G": 34}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row[0].row].height = 96

    dv = DataValidation(type="list", formula1='"' + ",".join(VERDICTS) + '"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Choose one of: " + ", ".join(VERDICTS)
    dv.errorTitle = "Not a valid verdict"
    dv.prompt = "correct / missing / incorrect / unsure"
    dv.promptTitle = "Verdict"
    ws.add_data_validation(dv)
    dv.add(f"F2:F{ws.max_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(head))}{ws.max_row}"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--annotators", default="Marko Petrov,Ema Pandilova")
    a = ap.parse_args()
    names = [n.strip() for n in a.annotators.split(",") if n.strip()]

    items = load_items()
    if not items:
        print("annotation_sheet.csv is empty; run export_human_validation first")
        return
    OUT.mkdir(parents=True, exist_ok=True)

    for i, name in enumerate(names):
        rows = list(items)
        # A distinct seeded order per annotator, so session drift cannot
        # correlate between them and inflate agreement.
        random.Random(1000 + i).shuffle(rows)
        s = slug(name)
        write_csv(OUT / f"annotation_{s}.csv", rows, name)
        write_xlsx(OUT / f"annotation_{s}.xlsx", rows, name)
        print(f"  {name}: {len(rows)} items -> annotation_{s}.xlsx and .csv")

    (OUT / "README.md").write_text(
        "# Human validation, how to do it\n\n"
        f"Two annotators, {len(items)} items each, the same items in different\n"
        "orders. Open the xlsx and use the dropdown in `human_verdict`, or edit\n"
        "the csv if you prefer Google Sheets. Both are the same items.\n\n"
        "```\n" + GUIDE + "```\n\n"
        "When you are done, send the file back and we run\n"
        "`python src/check_annotations.py` over both, which reports how often the\n"
        "two of you agree and how often each automated judge agrees with you.\n\n"
        "Roughly two to three hours each. It is the one figure that\n"
        "cannot be produced by a machine, and it is what lets us say the automated\n"
        "grader is unreliable rather than merely different.\n")
    print(f"\nwrote {OUT}")
    print("  README.md")


if __name__ == "__main__":
    main()
